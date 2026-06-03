import io
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st

# ── Constantes ─────────────────────────────────────────────────────────────────
C_HEADER_BG   = "1F4E79"
C_HEADER_FG   = "FFFFFF"
C_CORRECT     = "C6EFCE";  C_CORRECT_FG  = "276221"
C_ERROR       = "FFCCCC";  C_ERROR_FG    = "9C0006"
C_WARN        = "FFEB9C";  C_WARN_FG     = "9C6500"
C_ALT         = "F2F2F2"
C_ONLY_PROP   = "DAE3F3";  C_ONLY_PROP_FG  = "1F4E79"
C_ONLY_CONT   = "FCE4D6";  C_ONLY_CONT_FG  = "843C0C"
C_NOTE        = "FFFBEA";  C_NOTE_BORDER   = "E0C97A"

KEYWORDS = {
    '0':  ['saldo inicial'],
    '1':  ['canon', 'arrendamiento'],
    '2':  ['admon', 'administra', 'cuota'],
    '3':  ['iva'],
    '4':  ['retefte', 'fuente', 'reten'],
    '5':  ['ica'],
    '6':  ['reteiva', 'rete iva', 'iva'],
    '7':  ['epm', 'servicio', 'public', 'agua', 'tasa'],
    '8':  ['consignac'],
    '9':  ['iva', 'comis'],
    '10': ['4*1000', 'emergencia', 'gmf'],
    '11': ['seguro', 'poliza'],
    '12': ['cree', 'canon'],
    '14': ['iva', 'comis'],
    '15': ['iva', 'ph', 'admon'],
    '16': ['giro'],
    '18': ['derecho', 'contrato'],
    '20': ['comis'],
    '21': ['comis', 'admon'],
    '23': ['fuente', 'comi'],
    '24': ['ica', 'comis'],
    '25': ['iva', 'comis'],
    '26': ['cree', 'comis'],
    '28': ['bomberil', 'bomb', 'impuesto'],
    '29': ['aviso', 'impuesto'],
    '30': ['reparac', 'manteni', 'mtto', 'localiz'],
    '31': ['publicidad'],
    '32': ['predial'],
    '33': ['aseo', 'limpieza', 'anticipo'],
    '34': ['ingreso', 'canon'],
    '35': ['admon', 'ingreso'],
    '36': ['iva', 'ingreso'],
    '37': ['ajuste', 'reconoc', 'indemniz', 'otro'],
    '38': ['descont', 'ingreso', 'giro'],
    '39': ['multa', 'recargo'],
    '40': ['prorrateo', 'gasto'],
    '41': ['bomberil', 'comis'],
    '42': ['aviso', 'comis'],
}

# ── Catálogo Tipo Causa establecido (no cambia cada mes) ─────────────────────────
DEFAULT_TIPO_CAUSA = {
    '1':  'CANON ARRENDAMIENTO',
    '2':  'CUOTA DE ADMINISTRACION PH',
    '3':  'IVA DE ARRENDAMIENTO 19%',
    '4':  'RETENCION EN LA FUENTE CANON',
    '5':  'RETENCION DE ICA CANON',
    '6':  'RETENCION DE IVA ARRENDAMIENTO',
    '7':  'CUOTA SERVICIOS PUBLICOS',
    '8':  'CONSIGNACIONES',
    '9':  'RETENCION DE IVA COMISIO PROPIETARIO',
    '10': 'EMERGENCIA ECONOMICA 4*1000',
    '11': 'SEGURO PRIMAS',
    '12': 'RETENCION DE CREE CANON',
    '14': 'IVA 19% COMISION',
    '15': 'IVA 19% PH',
    '16': 'EGRESOS GIRO PROPIETARIOS',
    '18': 'DERECHOS DE CONTRATO CANON',
    '20': 'COMISION ARRENDAMIENTO',
    '21': 'COMISION ADMINISTRACION PH',
    '23': 'RETENCION EN LA FUENTE COMISIONES',
    '24': 'RETENCION DE ICA COMISION',
    '25': 'RETENCION DE IVA COMISION',
    '26': 'RETENCION DE CREE COMISION',
    '28': 'IMPUESTO BOMBERIL CANON',
    '29': 'IMPUESTO AVISOS CANON',
    '30': 'REPARACIONES LOCATIVAS',
    '31': 'PUBLICIDAD ARRENDAMIENTO',
    '32': 'IMPUESTO PREDIAL',
    '33': 'ASEO GASTO Y OTROS/ ANTICIPO PROPIETARIOS',
    '34': 'INGRESOS POR CANON',
    '35': 'INGRESOS POR ADMINISTRACION PH',
    '36': 'INGRESOS DE IVA ARRENDAMIENTO',
    '37': 'OTROS INGRESOS O INDEMNIZACIONES',
    '38': 'OTROS INGRESOS DESCONTABLES DE GIRO',
    '39': 'MULTAS Y RECARGOS',
    '40': 'PRORRATEO/ OTROS GASTOS',
    '41': 'IMPUESTO BOMBERILCOMISION',
    '42': 'IMPUESTO AVSO COMISION',
}

# ── Helpers ──────────────────────────────────────────────────────────────────
# Sufijos / palabras que delatan persona jurídica (empresa)
CORP_PAT = re.compile(
    r"\b(S\.?A\.?S|LTDA|SOCIEDAD|FUNDACI|ASOCIAC|INMOBILIARIA|CONSTRUCTORA|"
    r"INVERSIONES|GRUPO|COMPA[NÑ]I|\bCIA\b|\bE\.?U\b|CORPORACI|HOLDING|FONDO|"
    r"DESARROLLADOR|PROMOTORA|CONSORCIO|UNION TEMPORAL)\b"
)

def es_juridica(nit, nombre=""):
    """Identifica persona jurídica (empresa).
    Regla principal: NIT de 9 dígitos que empieza por 8 o 9.
    Regla secundaria: el nombre trae sufijo/palabra corporativa."""
    n = re.sub(r"\D", "", str(nit))
    if len(n) == 9 and n[:1] in ("8", "9"):
        return True
    if nombre and CORP_PAT.search(str(nombre).upper()):
        return True
    return False


def parse_money(v):
    """Convierte valores monetarios del auxiliar ('$977,182.00', '($ 940.00)')
    a entero. Negativos vienen entre paréntesis."""
    if pd.isna(v):
        return 0
    s = str(v).strip()
    if not s:
        return 0
    neg = "(" in s
    s = re.sub(r"[^\d.]", "", s)
    if not s or s == ".":
        return 0
    try:
        val = float(s)
    except ValueError:
        return 0
    return int(round(-val if neg else val))


# ── Clasificación de tipo de inmueble (a partir de la columna Detalle) ─────────
# El código del inmueble (AP/LC/PQV/PQM…) aparece junto al número de unidad al
# final del detalle, p. ej. "... UN MAJAGUA AP 708" o "... AGUACLARA LC3".
TIPO_INM_PAT = re.compile(
    r"(PQM|PQV|PQ|PV|APTO|AP|LOCAL|LC|OFC|C\.?U|BG|BODEGA|DEP[OÓ]SITO|LOTE)\s*\d",
    re.IGNORECASE)
TIPO_INM_MAP = {
    "AP": "AP", "APTO": "AP",
    "LC": "LC", "LOCAL": "LC",
    "PQV": "PQV", "PV": "PQV", "PQ": "PQV",
    "PQM": "PQM",
    "OFC": "OFC",
    "CU": "CU",
}
TIPO_INM_LABEL = {
    "AP":  "Apartamento (AP)",
    "LC":  "Local (LC)",
    "PQV": "Parqueadero vehículo (PQV)",
    "PQM": "Parqueadero moto (PQM)",
    "OFC": "Oficina (OFC)",
    "CU":  "Cuarto útil (C.U)",
    "OTRO": "Otro / sin clasificar",
}

def clasificar_inmueble(detalle):
    """Devuelve AP / LC / PQV / PQM u 'OTRO' según el código en el detalle."""
    s = str(detalle).upper()
    ms = TIPO_INM_PAT.findall(s)
    if not ms:
        return "OTRO"
    tok = ms[-1].upper().replace(".", "")
    return TIPO_INM_MAP.get(tok, "OTRO")


# ── Helpers Excel ──────────────────────────────────────────────────────────────
def _hdr(ws, row, ncols, bg=C_HEADER_BG, fg=C_HEADER_FG):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True, color=fg, name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="thin", color="000000"),
                             right=Side(style="thin", color="CCCCCC"))

def _row(ws, row, ncols, bg=None):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(vertical="center")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.border = Border(bottom=Side(style="thin", color="E0E0E0"),
                             right=Side(style="thin", color="E0E0E0"))

def _color(cell, bg, fg, bold=True):
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Arial", size=9, bold=bold, color=fg)

def _hdr_cell(ws, row, col, text, bg=C_HEADER_BG, fg=C_HEADER_FG):
    cell = ws.cell(row, col, text)
    cell.font = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=Side(style="thin", color="000000"),
                         right=Side(style="thin", color="CCCCCC"))

def _note_cell(ws, row, col):
    """Celda en blanco resaltada para que Contabilidad escriba observaciones."""
    cell = ws.cell(row, col)
    cell.fill = PatternFill("solid", fgColor=C_NOTE)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.font = Font(name="Arial", size=9, color="555555")
    cell.border = Border(top=Side(style="thin", color=C_NOTE_BORDER),
                         bottom=Side(style="thin", color=C_NOTE_BORDER),
                         left=Side(style="thin", color=C_NOTE_BORDER),
                         right=Side(style="thin", color=C_NOTE_BORDER))

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Lógica de conciliación ─────────────────────────────────────────────────────
def run_conciliacion(f_prop, f_cont, f_tc=None):
    # --- Propietarios activos ---
    df_p = pd.read_excel(f_prop, header=None)
    prop = df_p.iloc[1:]
    actv = prop[prop[12] == 'False'][[2, 13, 14]].copy()
    actv.columns = ["inm", "nombre_prop", "cedula_prop"]
    actv["inm"]         = actv["inm"].astype(str).str.strip()
    actv["nombre_prop"] = actv["nombre_prop"].astype(str).str.strip().str.upper()
    actv["cedula_prop"] = actv["cedula_prop"].astype(str).str.strip()
    prop_inm = actv.drop_duplicates("inm").set_index("inm")

    # --- Contabilidad (Auxiliar 28150501 — formato original) ---
    # El export puede traer o no la columna "Cuenta" al inicio (23 o 24 columnas),
    # por lo que se localizan las columnas por nombre de encabezado en lugar de
    # por posición fija.
    df_c = pd.read_excel(f_cont, header=None)
    hdr = df_c.iloc[0].astype(str).str.strip().str.lower()
    def _find_col(*names):
        for nm in names:
            for i, h in hdr.items():
                if h == nm.lower():
                    return i
        # coincidencia parcial como respaldo
        for nm in names:
            for i, h in hdr.items():
                if nm.lower() in h:
                    return i
        return None
    col_nit    = _find_col("nit")
    col_nombre = _find_col("nombre tercero")
    col_det    = _find_col("detalle")
    col_deb    = _find_col("debito", "débito")
    col_cred   = _find_col("credito", "crédito")
    col_inm    = _find_col("no. inm", "no inm", "inm")
    col_tc     = _find_col("tipocausa", "tipo causa")
    df_c = df_c.iloc[1:].copy()              # quita fila de encabezado
    # Mantiene solo filas de movimiento (descarta agrupaciones y totales)
    df_c = df_c[df_c[col_nombre].notna() & df_c[col_tc].notna()]
    cont = df_c.rename(columns={col_nit:"nit", col_nombre:"nombre_cont", col_det:"detalle",
                                col_deb:"debito", col_cred:"credito", col_inm:"inm", col_tc:"tipo_causa"})
    cont = cont[["nit","nombre_cont","detalle","debito","credito","inm","tipo_causa"]].copy()
    for col in ["nit", "nombre_cont", "detalle", "inm", "tipo_causa"]:
        cont[col] = cont[col].astype(str).str.strip()
    # TipoCausa puede venir como '1.0' al ser leído como float
    cont["tipo_causa"] = cont["tipo_causa"].str.replace(r"\.0$", "", regex=True)
    cont["inm"]        = cont["inm"].str.replace(r"\.0$", "", regex=True)
    cont["nombre_cont"] = cont["nombre_cont"].str.upper()
    cont["debito"]  = cont["debito"].map(parse_money)
    cont["credito"] = cont["credito"].map(parse_money)
    cont["juridica"] = [es_juridica(n, nm) for n, nm in zip(cont["nit"], cont["nombre_cont"])]
    cont = cont.reset_index(drop=True)

    cont_inm_owners = (
        cont[["inm","nit","nombre_cont"]]
        .drop_duplicates()
        .groupby("inm", group_keys=False)
        .apply(lambda df: df[["nit","nombre_cont"]].values.tolist())
        .to_dict()
    )

    # --- Tipo Causa ---
    if f_tc is not None:
        # El usuario subió un catálogo actualizado
        df_t = pd.read_excel(f_tc, header=None)
        tc = df_t.iloc[1:].copy()
        tc.columns = ["codigo", "concepto"]
        tc["codigo"]  = tc["codigo"].astype(str).str.strip()
        tc["concepto"]= tc["concepto"].astype(str).str.strip()
        tc_lookup = dict(zip(tc["codigo"], tc["concepto"]))
    else:
        # Se usa el catálogo establecido
        tc_lookup = dict(DEFAULT_TIPO_CAUSA)
    tc_lookup["0"] = "SALDO INICIAL"

    # ── Verificación 1 ──────────────────────────────────────────────────────────
    all_inm = sorted(
        set(prop_inm.index) | set(cont_inm_owners.keys()),
        key=lambda x: int(x) if x.isdigit() else float('inf')
    )
    v1_rows = []
    for inm in all_inm:
        in_p = inm in prop_inm.index
        in_c = inm in cont_inm_owners
        n_p  = prop_inm.loc[inm, "nombre_prop"] if in_p else "-"
        ced  = prop_inm.loc[inm, "cedula_prop"] if in_p else "-"

        if not in_c:
            v1_rows.append(dict(
                Inmueble=inm, Propietario_Prop=n_p, NIT_Prop=ced,
                Propietario_Cont="-", NIT_Cont="-",
                Estado="SOLO EN PROPIETARIOS",
                Alerta="Inmueble activo sin movimientos en Contabilidad"))
            continue

        owners = cont_inm_owners[inm]
        if not in_p:
            for nit_c, nom_c in owners:
                v1_rows.append(dict(
                    Inmueble=inm, Propietario_Prop="-", NIT_Prop="-",
                    Propietario_Cont=nom_c, NIT_Cont=nit_c,
                    Estado="SOLO EN CONTABILIDAD",
                    Alerta="Inmueble en Contabilidad sin propietario activo"))
            continue

        cont_nits = [r[0] for r in owners]
        if ced in cont_nits:
            nom_c = next(r[1] for r in owners if r[0] == ced)
            if " ".join(n_p.split()) == " ".join(nom_c.split()):
                estado, alerta = "CORRECTO", ""
            else:
                estado = "ALERTA — NOMBRE DIFIERE"
                alerta = f"NIT coincide pero nombre difiere. Cont: «{nom_c}»"
            v1_rows.append(dict(
                Inmueble=inm, Propietario_Prop=n_p, NIT_Prop=ced,
                Propietario_Cont=nom_c, NIT_Cont=ced,
                Estado=estado, Alerta=alerta))
            for nit_c, nom_c2 in owners:
                if nit_c != ced:
                    v1_rows.append(dict(
                        Inmueble=inm, Propietario_Prop="", NIT_Prop="",
                        Propietario_Cont=nom_c2, NIT_Cont=nit_c,
                        Estado="PROPIETARIO ADICIONAL",
                        Alerta="Contabilidad registra otro propietario en este inmueble"))
        else:
            cont_str = "; ".join(f"{n} ({ni})" for ni, n in owners)
            v1_rows.append(dict(
                Inmueble=inm, Propietario_Prop=n_p, NIT_Prop=ced,
                Propietario_Cont=owners[0][1], NIT_Cont=owners[0][0],
                Estado="ALERTA — NIT NO COINCIDE",
                Alerta=f"NIT {ced} no encontrado. Contabilidad tiene: {cont_str}"))

    df_v1 = pd.DataFrame(v1_rows)
    df_v1["Inmueble"] = pd.to_numeric(df_v1["Inmueble"], errors="coerce")
    df_v1 = df_v1.sort_values("Inmueble").reset_index(drop=True)

    # ── Verificación 2 ──────────────────────────────────────────────────────────
    def check_tc(tc_str, det_str):
        tc_s = str(tc_str).strip()
        det  = det_str.lower()
        conc = tc_lookup.get(tc_s)
        if not conc:
            return "CÓDIGO NO EXISTE", "-", f"TipoCausa '{tc_s}' no está en el catálogo"
        kws = KEYWORDS.get(tc_s, [])
        if not kws:
            return "SIN KEYWORDS", conc, "Sin palabras clave configuradas"
        if any(k in det for k in kws):
            return "CORRECTO", conc, ""
        return "REVISAR", conc, f"Keywords esperadas ({', '.join(kws)}) no encontradas"

    v2_rows = []
    for idx, row in cont.iterrows():
        est, conc, obs = check_tc(row["tipo_causa"], row["detalle"])
        v2_rows.append(dict(
            Fila=idx+2, Inmueble=row["inm"], NIT=row["nit"], Nombre=row["nombre_cont"],
            TipoCausa=row["tipo_causa"], Concepto=conc,
            Detalle=row["detalle"], Estado=est, Observacion=obs))

    df_v2 = pd.DataFrame(v2_rows)
    df_v2["Inmueble"] = pd.to_numeric(df_v2["Inmueble"], errors="coerce")
    df_v2 = df_v2.sort_values(["Inmueble","Fila"]).reset_index(drop=True)

    # ── Verificación 3 — IVA y Retenciones ─────────────────────────────────────
    df_v3 = cont[cont["tipo_causa"].isin(["3","4"])].copy()
    df_v3["inm_num"] = pd.to_numeric(df_v3["inm"], errors="coerce")
    df_v3 = df_v3.sort_values(["inm_num","tipo_causa"]).reset_index(drop=True)
    df_v3["concepto"] = df_v3["tipo_causa"].map(tc_lookup).fillna("-")
    df_v3["neto"] = df_v3["credito"] - df_v3["debito"]

    # --- TC1 (Canon): neto (débito − crédito) clasificado por tipo de inmueble ---
    t1 = cont[cont["tipo_causa"] == "1"].copy()
    t1["neto"] = t1["debito"] - t1["credito"]
    t1["tipo_inm"] = t1["detalle"].map(clasificar_inmueble)
    _orden = ["AP", "LC", "PQV", "PQM", "OFC", "CU", "OTRO"]
    tc1_by_type = []
    for tp in _orden:
        sub = t1[t1["tipo_inm"] == tp]
        if len(sub) == 0 and tp in ("OFC", "CU", "OTRO"):
            continue
        tc1_by_type.append(dict(
            tipo=tp, etiqueta=TIPO_INM_LABEL[tp],
            cantidad=int(len(sub)), neto=int(sub["neto"].sum())))
    tc1_neto_total = int(t1["neto"].sum())
    tc1_cant_total = int(len(t1))

    # Detalle de los canon (TC1) que NO se pudieron clasificar (tipo OTRO)
    tc1_sin_clasif = t1[t1["tipo_inm"] == "OTRO"].copy()
    tc1_sin_clasif["inm_num"] = pd.to_numeric(tc1_sin_clasif["inm"], errors="coerce")
    tc1_sin_clasif = tc1_sin_clasif.sort_values("inm_num")[
        ["inm", "nit", "nombre_cont", "detalle", "neto"]
    ].reset_index(drop=True)

    # --- Netos TC3 (IVA arrend.) y TC4 (Retefuente canon): débito − crédito ---
    _t3 = cont[cont["tipo_causa"] == "3"]
    _t4 = cont[cont["tipo_causa"] == "4"]
    tc3_neto = int(_t3["debito"].sum() - _t3["credito"].sum())
    tc4_neto = int(_t4["debito"].sum() - _t4["credito"].sum())

    v3_extra = dict(
        tc1_by_type=tc1_by_type, tc1_neto_total=tc1_neto_total,
        tc1_cant_total=tc1_cant_total, tc3_neto=tc3_neto, tc4_neto=tc4_neto,
        tc1_sin_clasif=tc1_sin_clasif,
    )

    # ── Verificaciones cruzadas IVA y Retefuente ─────────────────────────────
    # Naturaleza: TC20/TC14 son débitos, TC23 es crédito. Se comparan magnitudes.
    def base_tc(tc_str, jur=None):
        sub = cont[cont["tipo_causa"] == tc_str]
        if jur is True:  sub = sub[sub["juridica"]]
        if jur is False: sub = sub[~sub["juridica"]]
        return abs(int(sub["credito"].sum() - sub["debito"].sum()))

    val_20_all = base_tc("20")            # TC20 total (naturales + jurídicas)
    val_20_jur = base_tc("20", True)      # TC20 solo jurídicas
    val_20_nat = base_tc("20", False)     # TC20 solo naturales
    val_14     = base_tc("14")            # IVA comisión (todos)
    val_23     = base_tc("23")            # Retefuente comisión (solo jurídicas)

    # El impuesto se calcula y redondea por transacción (como en contabilidad),
    # no sobre el agregado, para no introducir errores de redondeo.
    t20  = cont[cont["tipo_causa"] == "20"]
    t20j = t20[t20["juridica"]]
    _base20  = (t20["credito"]  - t20["debito"]).abs()
    _base20j = (t20j["credito"] - t20j["debito"]).abs()
    iva_calc   = int((_base20  * 0.19).round().sum())   # TODO el TC20 × 19% debe = TC14
    rete_calc  = int((_base20j * 0.11).round().sum())   # solo jurídicas TC20 × 11% debe = TC23
    dif_iva    = iva_calc  - val_14
    dif_rete   = rete_calc - val_23

    # Jurídicas con TC20 pero SIN TC23 → error
    nom_by_nit  = cont.groupby("nit")["nombre_cont"].first()
    nits_20_jur = set(cont[(cont["tipo_causa"]=="20") & cont["juridica"]]["nit"])
    nits_23_jur = set(cont[(cont["tipo_causa"]=="23") & cont["juridica"]]["nit"])
    jur_sin_tc23 = [(n, nom_by_nit.get(n, "-")) for n in sorted(nits_20_jur - nits_23_jur)]

    crosscheck = dict(
        val_20_all=val_20_all, val_20_jur=val_20_jur, val_20_nat=val_20_nat,
        val_14=val_14, val_23=val_23,
        iva_calc=iva_calc, rete_calc=rete_calc,
        dif_iva=dif_iva, dif_rete=dif_rete,
        ok_iva=(dif_iva==0), ok_rete=(dif_rete==0),
        jur_sin_tc23=jur_sin_tc23,
        n_juridicas=cont[cont["juridica"]]["nit"].nunique(),
        n_naturales=cont[~cont["juridica"]]["nit"].nunique(),
        # claves de compatibilidad
        neto_20=val_20_all, neto_14=val_14, neto_23=val_23,
        # Verificación 3 ampliada (TC1 por tipo de inmueble, netos TC3/TC4)
        **v3_extra,
    )

    # ── Análisis de diferencias por inmueble ─────────────────────────────────
    _piv = cont[cont["tipo_causa"].isin(["14","20","23"])].groupby(["inm","tipo_causa"]).apply(
        lambda x: abs(int(x["credito"].sum() - x["debito"].sum())), include_groups=False
    ).unstack(fill_value=0)
    _piv.columns.name = None
    for _c in ["14","20","23"]:
        if _c not in _piv.columns: _piv[_c] = 0
    _piv = _piv[["14","20","23"]].copy()
    _piv.columns = ["neto_14","neto_20","neto_23"]
    _piv["iva_calc"]  = (_piv["neto_20"] * 0.19).round().astype(int)
    _piv["rete_calc"] = (_piv["neto_20"] * 0.11).round().astype(int)
    _piv["dif_iva"]   = _piv["iva_calc"]  - _piv["neto_14"]
    _piv["dif_rete"]  = _piv["rete_calc"] - _piv["neto_23"]
    _piv["inm_num"]   = pd.to_numeric(_piv.index, errors="coerce")
    _nom = cont[cont["tipo_causa"].isin(["14","20","23"])].groupby("inm")["nombre_cont"].first()
    _jur = cont[cont["tipo_causa"].isin(["14","20","23"])].groupby("inm")["juridica"].max()
    _piv["nombre"]   = _piv.index.map(_nom)
    _piv["juridica"] = _piv.index.map(_jur).fillna(False)

    def _clr(row):
        if not row["juridica"]:        return "NATURAL (sin retención)"  # OK, no aplica TC23
        if row["neto_23"] == 0:        return "JURÍDICA SIN TC23"        # ERROR
        if row["dif_rete"] == 0:       return "OK"
        return "DIFERENCIA"

    _piv["tipo_dif_rete"] = _piv.apply(_clr, axis=1)
    _piv["tipo_dif_iva"]  = _piv.apply(
        lambda r: "OK" if r["dif_iva"]==0 else ("SIN TC14" if r["neto_14"]==0 else "DIFERENCIA PARCIAL"),
        axis=1)

    df_iva_dif  = _piv[_piv["dif_iva"]!=0].sort_values("inm_num").reset_index()
    # Solo errores reales de retefuente (jurídicas): sin TC23 o con diferencia
    df_rete_dif = _piv[_piv["tipo_dif_rete"].isin(["JURÍDICA SIN TC23","DIFERENCIA"])]\
                      .sort_values("inm_num").reset_index()
    resumen_rete = _piv.groupby("tipo_dif_rete").agg(
        inmuebles=("dif_rete","count"), dif_total=("dif_rete","sum")
    ).reset_index()

    analisis = dict(
        df_iva_dif=df_iva_dif, df_rete_dif=df_rete_dif, resumen_rete=resumen_rete,
        n_jur_sin_tc23=len(_piv[_piv["tipo_dif_rete"]=="JURÍDICA SIN TC23"]),
        n_jur_dif=len(_piv[_piv["tipo_dif_rete"]=="DIFERENCIA"]),
        n_sin_tc14=len(_piv[_piv["tipo_dif_iva"]=="SIN TC14"]),
        sum_jur_sin_tc23=int(_piv[_piv["tipo_dif_rete"]=="JURÍDICA SIN TC23"]["rete_calc"].sum()),
    )

    return df_v1, df_v2, df_v3, crosscheck, analisis, prop_inm, cont_inm_owners


# ── Generar Excel ──────────────────────────────────────────────────────────────
def build_excel(df_v1, df_v2, df_v3, crosscheck, analisis):
    wb  = openpyxl.Workbook()

    # ── Resumen ──────────────────────────────────────────────────────────────
    ws  = wb.active; ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "REPORTE DE CONCILIACIÓN RM2 — CORTE ABRIL 2026"
    t.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A2:F2")
    ws["A2"].value = "Solo propietarios activos (Ex Prop = False)"
    ws["A2"].font  = Font(name="Arial", size=9, italic=True, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")

    FMT_NUM   = '#,##0'
    total_inm = df_v1["Inmueble"].nunique()
    correctos = len(df_v1[df_v1["Estado"]=="CORRECTO"])
    al_nombre = len(df_v1[df_v1["Estado"]=="ALERTA — NOMBRE DIFIERE"])
    al_nit    = len(df_v1[df_v1["Estado"]=="ALERTA — NIT NO COINCIDE"])
    solo_p    = len(df_v1[df_v1["Estado"]=="SOLO EN PROPIETARIOS"])
    solo_c    = len(df_v1[df_v1["Estado"]=="SOLO EN CONTABILIDAD"])
    adicional = len(df_v1[df_v1["Estado"]=="PROPIETARIO ADICIONAL"])
    tot_v2    = len(df_v2)
    cor_v2    = len(df_v2[df_v2["Estado"]=="CORRECTO"])
    rev_v2    = len(df_v2[df_v2["Estado"]=="REVISAR"])
    sin_v2    = len(df_v2[df_v2["Estado"]=="CÓDIGO NO EXISTE"])
    tot_v3    = len(df_v3)
    iva_v3    = len(df_v3[df_v3["tipo_causa"]=="3"])
    rete_v3   = len(df_v3[df_v3["tipo_causa"]=="4"])
    deb_v3    = int(df_v3["debito"].sum())
    cred_v3   = int(df_v3["credito"].sum())
    neto_v3   = cred_v3 - deb_v3
    cc        = crosscheck

    def section(start, title, rows):
        ws.cell(start, 1, title).font = Font(name="Arial", size=11, bold=True, color=C_HEADER_BG)
        ws.cell(start+1, 1, "Métrica"); ws.cell(start+1, 2, "Valor")
        _hdr(ws, start+1, 2)
        _hdr_cell(ws, start+1, 4, "Observaciones (Contabilidad)")
        for off, (lbl, val, bg, fg) in enumerate(rows):
            r = start + 2 + off
            ws.cell(r, 1, lbl).font = Font(name="Arial", size=9)
            cv = ws.cell(r, 2, val); cv.alignment = Alignment(horizontal="center")
            _row(ws, r, 2, C_ALT if r%2==0 else None)
            if bg: _color(cv, bg, fg)
            _note_cell(ws, r, 4)
        return start + 2 + len(rows) + 2

    def v3_section(start, title, rows):
        """Sección con columnas Concepto | Cantidad | Neto (déb−créd)."""
        ws.cell(start, 1, title).font = Font(name="Arial", size=11, bold=True, color=C_HEADER_BG)
        for c, h in enumerate(["Concepto", "Cantidad", "Neto (déb − créd)"], 1):
            ws.cell(start+1, c, h)
        _hdr(ws, start+1, 3)
        _hdr_cell(ws, start+1, 4, "Observaciones (Contabilidad)")
        for off, (lbl, cant, neto, bg, fg) in enumerate(rows):
            r = start + 2 + off
            ws.cell(r, 1, lbl).font = Font(name="Arial", size=9, bold=(bg is not None))
            cc1 = ws.cell(r, 2, cant); cc1.alignment = Alignment(horizontal="center")
            cc2 = ws.cell(r, 3, neto); cc2.alignment = Alignment(horizontal="right")
            if neto != "": cc2.number_format = FMT_NUM
            _row(ws, r, 3, C_ALT if r%2==0 else None)
            if bg:
                for c in [1, 2, 3]: _color(ws.cell(r, c), bg, fg)
            _note_cell(ws, r, 4)
        return start + 2 + len(rows) + 2

    def crosscheck_section(start, title, rows):
        t2 = ws.cell(start, 1, title)
        t2.font = Font(name="Arial", size=11, bold=True, color=C_HEADER_BG)
        ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=4)
        for c, h in enumerate(["Concepto","Valor",""], 1):
            ws.cell(start+1, c, h)
        _hdr(ws, start+1, 3)
        _hdr_cell(ws, start+1, 4, "Observaciones (Contabilidad)")
        for off, (lbl, val, num_fmt, bg, fg) in enumerate(rows):
            r = start + 2 + off
            ws.cell(r, 1, lbl).font = Font(name="Arial", size=9)
            cv = ws.cell(r, 2, val)
            cv.font = Font(name="Arial", size=9, bold=(bg is not None))
            cv.alignment = Alignment(horizontal="right")
            if num_fmt: cv.number_format = num_fmt
            _row(ws, r, 3, C_ALT if r%2==0 else None)
            if bg:
                for c in [1, 2]:
                    cell = ws.cell(r, c)
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(name="Arial", size=9, bold=True, color=fg)
            _note_cell(ws, r, 4)
        return start + 2 + len(rows) + 2

    r0 = section(4, "VERIFICACIÓN 1 — PROPIETARIOS POR INMUEBLE", [
        ("Total inmuebles verificados",                    total_inm, None,        "000000"),
        ("✓ Correctos (nombre e ID coinciden)",            correctos, C_CORRECT,   C_CORRECT_FG),
        ("⚠ Alerta — nombre difiere",                     al_nombre, C_WARN,      C_WARN_FG),
        ("✗ Alerta — NIT no coincide",                    al_nit,    C_ERROR,     C_ERROR_FG),
        ("→ Solo en Propietarios (sin movimientos)",       solo_p,    C_ONLY_PROP, C_ONLY_PROP_FG),
        ("→ Solo en Contabilidad (sin propietario activo)",solo_c,    C_ONLY_CONT, C_ONLY_CONT_FG),
        ("ℹ Propietarios adicionales en Contabilidad",    adicional, None,        "888888"),
    ])
    r0 = section(r0, "VERIFICACIÓN 2 — TIPO CAUSA", [
        ("Total filas verificadas",                        tot_v2, None,      "000000"),
        ("✓ Correctos",                                    cor_v2, C_CORRECT, C_CORRECT_FG),
        ("⚠ Para revisar",                                rev_v2, C_WARN,    C_WARN_FG),
        ("✗ Código no existe en catálogo",                 sin_v2, C_ERROR,   C_ERROR_FG),
    ])
    # ── Verificación 3 — Canon (TC1) por tipo de inmueble + netos TC3/TC4 ──────
    _tc1_rows = [
        (d["etiqueta"], d["cantidad"], d["neto"], None, "000000")
        for d in cc["tc1_by_type"]
    ]
    _tc1_rows.append(("TOTAL CANON (TC 1)", cc["tc1_cant_total"], cc["tc1_neto_total"],
                      C_HEADER_BG, "FFFFFF"))
    r0 = v3_section(r0, "VERIFICACIÓN 3A — CANON (TC 1) POR TIPO DE INMUEBLE", _tc1_rows)
    r0 = v3_section(r0, "VERIFICACIÓN 3B — IVA Y RETENCIONES (TC 3 y TC 4)", [
        ("TC 3 — IVA Arrendamiento 19%",          iva_v3,  cc["tc3_neto"], None, "000000"),
        ("TC 4 — Retención en la Fuente Canon",   rete_v3, cc["tc4_neto"], None, "000000"),
        ("Total registros (TC 3 + TC 4)",         tot_v3,  "",             None, "000000"),
    ])

    # ── Verificación 3C — Canon (TC1) SIN clasificación ───────────────────────
    df_sc = cc["tc1_sin_clasif"]
    ws.cell(r0, 1, "VERIFICACIÓN 3C — CANON (TC 1) SIN CLASIFICACIÓN").font = Font(
        name="Arial", size=11, bold=True, color=C_HEADER_BG)
    sc_cols = ["No. Inm", "NIT", "Nombre Tercero", "Detalle", "Neto (déb − créd)"]
    for c, h in enumerate(sc_cols, 1):
        _hdr_cell(ws, r0+1, c, h)
    if len(df_sc) == 0:
        ws.merge_cells(start_row=r0+2, start_column=1, end_row=r0+2, end_column=5)
        ok = ws.cell(r0+2, 1, "✓ Todos los canon (TC 1) quedaron clasificados")
        ok.font = Font(name="Arial", size=9, color=C_CORRECT_FG)
        ok.fill = PatternFill("solid", fgColor=C_CORRECT)
        r0 = r0 + 2 + 2
    else:
        for ri, rw in df_sc.iterrows():
            r = r0 + 2 + ri
            for ci, v in enumerate([rw["inm"], rw["nit"], rw["nombre_cont"], rw["detalle"], rw["neto"]], 1):
                ws.cell(r, ci, v).font = Font(name="Arial", size=9)
            ws.cell(r, 5).number_format = FMT_NUM
            ws.cell(r, 5).alignment = Alignment(horizontal="right")
            _row(ws, r, 5, C_ALT if r%2==0 else None)
            _color(ws.cell(r, 1), C_WARN, C_WARN_FG)
        r0 = r0 + 2 + len(df_sc) + 2

    r0 = crosscheck_section(r0, "VERIFICACIÓN CRUZADA IVA — TC20 (total) × 19% debe = TC14", [
        ("Valor TC 20 Comisión — TOTAL (naturales + jurídicas)", cc["val_20_all"], FMT_NUM, None, "000000"),
        ("   · TC 20 personas naturales",             cc["val_20_nat"], FMT_NUM, None,     "000000"),
        ("   · TC 20 personas jurídicas",             cc["val_20_jur"], FMT_NUM, None,     "000000"),
        ("× IVA 19% → valor calculado",               cc["iva_calc"],  FMT_NUM, None,     "000000"),
        ("Valor TC 14 (IVA 19% Comisión) — real",     cc["val_14"],    FMT_NUM, None,     "000000"),
        ("Diferencia (Calculado − Real)",              cc["dif_iva"],   FMT_NUM,
         None if cc["ok_iva"] else C_ERROR, C_CORRECT_FG if cc["ok_iva"] else C_ERROR_FG),
        ("Estado",
         "✓ CUADRA" if cc["ok_iva"] else f"✗ DIFIERE en {cc['dif_iva']:,}",
         None,
         C_CORRECT if cc["ok_iva"] else C_ERROR, C_CORRECT_FG if cc["ok_iva"] else C_ERROR_FG),
    ])
    r0 = crosscheck_section(r0, "VERIFICACIÓN CRUZADA RETEFUENTE — TC20 jurídicas × 11% debe = TC23", [
        ("Valor TC 20 Comisión — solo JURÍDICAS",     cc["val_20_jur"], FMT_NUM, None,     "000000"),
        ("× Retefuente 11% → valor calculado",        cc["rete_calc"],  FMT_NUM, None,     "000000"),
        ("Valor TC 23 (Retención Fuente Comisión) — real", cc["val_23"], FMT_NUM, None,   "000000"),
        ("Diferencia (Calculado − Real)",              cc["dif_rete"],   FMT_NUM,
         None if cc["ok_rete"] else C_ERROR, C_CORRECT_FG if cc["ok_rete"] else C_ERROR_FG),
        ("Estado",
         "✓ CUADRA" if cc["ok_rete"] else f"✗ DIFIERE en {cc['dif_rete']:,}",
         None,
         C_CORRECT if cc["ok_rete"] else C_ERROR, C_CORRECT_FG if cc["ok_rete"] else C_ERROR_FG),
        ("Nota: personas naturales no tienen TC23 (no es error)", "", None, None, "808080"),
    ])
    # ── Diagnóstico ──────────────────────────────────────────────────────────
    an = analisis
    _dt = r0           # fila título del diagnóstico
    ws.cell(_dt, 1, "DIAGNÓSTICO DE DIFERENCIAS").font = Font(
        name="Arial", size=11, bold=True, color=C_HEADER_BG)
    ws.cell(_dt+1, 1, "Diagnóstico"); ws.cell(_dt+1, 2, "Cantidad")
    _hdr(ws, _dt+1, 2)
    _hdr_cell(ws, _dt+1, 4, "Observaciones (Contabilidad)")
    diag_rows = [
        ("IVA — Inmuebles sin TC14",                              an["n_sin_tc14"],     C_ERROR if an["n_sin_tc14"]     else C_CORRECT, C_ERROR_FG if an["n_sin_tc14"]     else C_CORRECT_FG),
        ("Retefuente — Inmuebles JURÍDICOS sin TC23 (error)",     an["n_jur_sin_tc23"], C_ERROR if an["n_jur_sin_tc23"] else C_CORRECT, C_ERROR_FG if an["n_jur_sin_tc23"] else C_CORRECT_FG),
        ("Retefuente — Inmuebles JURÍDICOS con diferencia",       an["n_jur_dif"],      C_WARN  if an["n_jur_dif"]      else C_CORRECT, C_WARN_FG  if an["n_jur_dif"]      else C_CORRECT_FG),
        ("Retefuente faltante acumulada (jurídicas sin TC23)",    an["sum_jur_sin_tc23"], C_ERROR if an["sum_jur_sin_tc23"] else C_CORRECT, C_ERROR_FG if an["sum_jur_sin_tc23"] else C_CORRECT_FG),
    ]
    for off, (lbl, val, bg, fg) in enumerate(diag_rows):
        r = _dt + 2 + off
        ws.cell(r, 1, lbl).font = Font(name="Arial", size=9)
        cv = ws.cell(r, 2, val); cv.alignment = Alignment(horizontal="center")
        cv.number_format = FMT_NUM
        _row(ws, r, 2, C_ALT if r%2==0 else None)
        _color(cv, bg, fg)
        _note_cell(ws, r, 4)

    _widths(ws, [52, 16, 20, 45, 18])

    # ── Ver4 Análisis de Diferencias ─────────────────────────────────────────
    ws4 = wb.create_sheet("Ver4 Análisis Diferencias")
    ws4.sheet_view.showGridLines = False

    def v4t(row, text, ncols):
        ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        t2 = ws4.cell(row, 1, text)
        t2.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        t2.fill = PatternFill("solid", fgColor=C_HEADER_BG)
        t2.alignment = Alignment(horizontal="center", vertical="center")
        ws4.row_dimensions[row].height = 22

    FMT_NUM2 = '#,##0'
    dcols = ["No. Inm","Nombre Tercero","Neto TC20","Calculado","Neto Real","Diferencia","Diagnóstico","Observaciones (Contabilidad)"]
    NC = len(dcols)  # número de columnas de las tablas de detalle

    # Sección IVA
    v4t(1, "DIFERENCIAS IVA — Inmuebles con TC20 sin TC14", NC)
    for c, h in enumerate(dcols, 1): ws4.cell(2, c, h)
    _hdr(ws4, 2, NC)
    df_id = an["df_iva_dif"]
    if len(df_id) == 0:
        ws4.merge_cells(start_row=3, start_column=1, end_row=3, end_column=NC)
        c2 = ws4.cell(3, 1, "✓ Sin diferencias de IVA")
        c2.font = Font(name="Arial", size=9, color=C_CORRECT_FG)
        c2.fill = PatternFill("solid", fgColor=C_CORRECT)
    else:
        for ri, rw in df_id.iterrows():
            r = ri + 3
            for ci, v in enumerate([rw["inm_num"],rw["nombre"],rw["neto_20"],rw["iva_calc"],rw["neto_14"],rw["dif_iva"],rw["tipo_dif_iva"]], 1):
                ws4.cell(r, ci, v)
            _row(ws4, r, NC, C_ALT if ri%2==0 else None)
            for ci in [3,4,5,6]: ws4.cell(r,ci).number_format=FMT_NUM2; ws4.cell(r,ci).alignment=Alignment(horizontal="right")
            _color(ws4.cell(r,7), C_ERROR, C_ERROR_FG)
            _note_cell(ws4, r, NC)

    sep = len(df_id) + 5
    # Resumen retefuente
    v4t(sep, "RETEFUENTE — Resumen por categoría", NC)
    for c, h in enumerate(["Categoría","Inmuebles","Diferencia","Interpretación"], 1): ws4.cell(sep+1, c, h)
    _hdr(ws4, sep+1, 4)
    cats = {
        "JURÍDICA SIN TC23":(C_ERROR,C_ERROR_FG,"ERROR: empresa sin retención registrada"),
        "DIFERENCIA":(C_WARN,C_WARN_FG,"Empresa con TC23 distinto al 11% del TC20"),
        "NATURAL (sin retención)":(C_CORRECT,C_CORRECT_FG,"OK: persona natural, no aplica TC23"),
        "OK":(C_CORRECT,C_CORRECT_FG,"OK: TC23 = 11% del TC20"),
    }
    _rr = sep + 2
    for _, rw in an["resumen_rete"].iterrows():
        cat = rw["tipo_dif_rete"]
        bg2, fg2, interp = cats.get(cat, (None,"000000",""))
        ws4.cell(_rr,1,cat); ws4.cell(_rr,2,int(rw["inmuebles"])).alignment=Alignment(horizontal="center")
        ws4.cell(_rr,3,int(rw["dif_total"])).number_format=FMT_NUM2; ws4.cell(_rr,3).alignment=Alignment(horizontal="right")
        ws4.cell(_rr,4,interp); _row(ws4,_rr,4,C_ALT if _rr%2==0 else None)
        if bg2:
            for ci in [1,2,3]: _color(ws4.cell(_rr,ci),bg2,fg2)
        _rr += 1

    sep2 = _rr + 2
    # Jurídicas sin TC23 (ERROR)
    v4t(sep2, "DETALLE — Inmuebles JURÍDICOS SIN TC23 (retefuente no registrada — ERROR)", NC)
    for c, h in enumerate(dcols, 1): ws4.cell(sep2+1, c, h)
    _hdr(ws4, sep2+1, NC)
    df_sc = an["df_rete_dif"][an["df_rete_dif"]["tipo_dif_rete"]=="JURÍDICA SIN TC23"].reset_index(drop=True)
    if len(df_sc) == 0:
        ws4.merge_cells(start_row=sep2+2, start_column=1, end_row=sep2+2, end_column=NC)
        c2 = ws4.cell(sep2+2, 1, "✓ Todas las personas jurídicas tienen TC23")
        c2.font = Font(name="Arial", size=9, color=C_CORRECT_FG)
        c2.fill = PatternFill("solid", fgColor=C_CORRECT)
    for ri, rw in df_sc.iterrows():
        r = sep2 + 2 + ri
        for ci, v in enumerate([rw["inm_num"],rw["nombre"],rw["neto_20"],rw["rete_calc"],rw["neto_23"],rw["dif_rete"],"TC23 NO REGISTRADO"], 1):
            ws4.cell(r, ci, v)
        _row(ws4, r, NC, C_ALT if ri%2==0 else None)
        for ci in [3,4,5,6]: ws4.cell(r,ci).number_format=FMT_NUM2; ws4.cell(r,ci).alignment=Alignment(horizontal="right")
        _color(ws4.cell(r,7), C_ERROR, C_ERROR_FG)
        _note_cell(ws4, r, NC)

    sep3 = sep2 + 2 + max(len(df_sc),1) + 2
    # Jurídicas con diferencia
    v4t(sep3, "DETALLE — Inmuebles JURÍDICOS con DIFERENCIA (TC23 ≠ 11% del TC20)", NC)
    for c, h in enumerate(dcols, 1): ws4.cell(sep3+1, c, h)
    _hdr(ws4, sep3+1, NC)
    df_si = an["df_rete_dif"][an["df_rete_dif"]["tipo_dif_rete"]=="DIFERENCIA"].reset_index(drop=True)
    if len(df_si) == 0:
        ws4.merge_cells(start_row=sep3+2, start_column=1, end_row=sep3+2, end_column=NC)
        c3 = ws4.cell(sep3+2, 1, "✓ Sin diferencias de retefuente en personas jurídicas")
        c3.font = Font(name="Arial", size=9, color=C_CORRECT_FG)
        c3.fill = PatternFill("solid", fgColor=C_CORRECT)
    for ri, rw in df_si.iterrows():
        r = sep3 + 2 + ri
        for ci, v in enumerate([rw["inm_num"],rw["nombre"],rw["neto_20"],rw["rete_calc"],rw["neto_23"],rw["dif_rete"],"DIFERENCIA"], 1):
            ws4.cell(r, ci, v)
        _row(ws4, r, NC, C_ALT if ri%2==0 else None)
        for ci in [3,4,5,6]: ws4.cell(r,ci).number_format=FMT_NUM2; ws4.cell(r,ci).alignment=Alignment(horizontal="right")
        _color(ws4.cell(r,7), C_WARN, C_WARN_FG)
        _note_cell(ws4, r, NC)

    _widths(ws4, [12, 35, 16, 20, 16, 16, 30, 45])
    ws4.freeze_panes = "A3"

    # ── Ver1 Propietarios ─────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Ver1 Propietarios")
    ws1.sheet_view.showGridLines = False
    ws1.merge_cells("A1:G1")
    t = ws1["A1"]
    t.value = "VERIFICACIÓN 1 — Propietarios por Inmueble (solo activos)"
    t.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 25

    cols1 = ["No. Inmueble","Propietario (Propietarios)","Cédula/NIT (Prop.)",
             "Propietario (Contabilidad)","NIT (Contabilidad)","Estado","⚠ Alerta"]
    src1  = ["Inmueble","Propietario_Prop","NIT_Prop",
             "Propietario_Cont","NIT_Cont","Estado","Alerta"]
    for c, h in enumerate(cols1, 1): ws1.cell(2, c, h)
    _hdr(ws1, 2, len(cols1))

    ec_col = 6
    al_col = 7
    for ri, row in df_v1.iterrows():
        r = ri + 3
        for ci, s in enumerate(src1, 1): ws1.cell(r, ci, row[s])
        _row(ws1, r, len(cols1), C_ALT if ri%2==0 else None)
        ec = ws1.cell(r, ec_col)
        st = row["Estado"]
        if st == "CORRECTO":
            _color(ec, C_CORRECT, C_CORRECT_FG)
        elif "ALERTA" in st:
            _color(ec, C_ERROR, C_ERROR_FG)
            _color(ws1.cell(r, al_col), C_ERROR, C_ERROR_FG, bold=False)
        elif st == "SOLO EN PROPIETARIOS":
            _color(ec, C_ONLY_PROP, C_ONLY_PROP_FG)
        elif st == "SOLO EN CONTABILIDAD":
            _color(ec, C_ONLY_CONT, C_ONLY_CONT_FG)
        elif st == "PROPIETARIO ADICIONAL":
            _color(ec, C_WARN, C_WARN_FG)

    _widths(ws1, [13,33,18,33,18,26,55])
    ws1.freeze_panes = "A3"

    # ── Ver2 TipoCausa ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Ver2 TipoCausa")
    ws2.sheet_view.showGridLines = False
    ws2.merge_cells("A1:I1")
    t = ws2["A1"]
    t.value = "VERIFICACIÓN 2 — Tipo Causa vs Detalle (ordenado por No. Inmueble)"
    t.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 25

    cols2 = ["Fila","No. Inmueble","NIT","Nombre","TipoCausa",
             "Concepto (Catálogo)","Detalle Contabilidad","Estado","Observación"]
    src2  = ["Fila","Inmueble","NIT","Nombre","TipoCausa",
             "Concepto","Detalle","Estado","Observacion"]
    for c, h in enumerate(cols2, 1): ws2.cell(2, c, h)
    _hdr(ws2, 2, len(cols2))

    ec2 = cols2.index("Estado") + 1
    for ri, row in df_v2.iterrows():
        r = ri + 3
        for ci, s in enumerate(src2, 1): ws2.cell(r, ci, row[s])
        _row(ws2, r, len(cols2), C_ALT if ri%2==0 else None)
        ec = ws2.cell(r, ec2)
        if row["Estado"] == "CORRECTO":
            _color(ec, C_CORRECT, C_CORRECT_FG)
        elif row["Estado"] == "REVISAR":
            _color(ec, C_WARN, C_WARN_FG)
        elif row["Estado"] in ("CÓDIGO NO EXISTE","SIN KEYWORDS"):
            _color(ec, C_ERROR, C_ERROR_FG)

    _widths(ws2, [6,13,14,30,11,30,60,15,45])
    ws2.freeze_panes = "A3"

    # ── Ver3 IVA y Retenciones ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Ver3 IVA y Retenciones")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:H1")
    t = ws3["A1"]
    t.value = "VERIFICACIÓN 3 — IVA (TC 3) y Retención en la Fuente (TC 4)"
    t.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 25

    cols3 = ["No. Inmueble","Nombre Tercero","TC","Concepto","Detalle","Débito","Crédito","Neto (Cred - Deb)"]
    for c, h in enumerate(cols3, 1): ws3.cell(2, c, h)
    _hdr(ws3, 2, len(cols3))

    C_IVA  = "DEEAF1"
    C_RETE = "EBF0DE"

    for ri, row in df_v3.iterrows():
        r = ri + 3
        neto = int(row["credito"]) - int(row["debito"])
        vals = [row["inm_num"], row["nombre_cont"], row["tipo_causa"],
                row["concepto"], row["detalle"], int(row["debito"]), int(row["credito"]), neto]
        for ci, val in enumerate(vals, 1): ws3.cell(r, ci, val)
        bg = C_IVA if row["tipo_causa"] == "3" else C_RETE
        _row(ws3, r, len(cols3), bg)
        for ci in [1, 6, 7, 8]:
            ws3.cell(r, ci).alignment = Alignment(horizontal="right")
        if neto < 0:
            ws3.cell(r, 8).font = Font(name="Arial", size=9, bold=True, color=C_ERROR_FG)

    r_tot = len(df_v3) + 3
    ws3.cell(r_tot, 5, "TOTAL").font = Font(name="Arial", size=9, bold=True)
    tot_deb  = int(df_v3["debito"].sum())
    tot_cred = int(df_v3["credito"].sum())
    tot_neto = tot_cred - tot_deb
    for ci, val in [(6, tot_deb), (7, tot_cred), (8, tot_neto)]:
        cell = ws3.cell(r_tot, ci, val)
        cell.font = Font(name="Arial", size=9, bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.alignment = Alignment(horizontal="right")
        cell.border = Border(top=Side(style="thin", color="000000"))

    _widths(ws3, [14,32,5,32,65,14,14,18])
    ws3.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf




# ==============================================================================
# MÓDULO: PLANTILLA CONTAI (Actas Fondo)
# ==============================================================================

NIT_FONDO        = 901155077
CENTRO_COSTO     = "0501"
CUENTA_CONTRA    = 22050501
CUENTA_TOTAL_ACTA = 13050507   # contrapartida de la fila "Total acta mes"

C_HEADER_BG = "1F4E79"
C_HEADER_FG = "FFFFFF"
C_ALT       = "F2F2F2"

CUENTAS = {
    1:  {"gravado": 41550502, "no_gravado": 41550501},
    2:  61053502,
    3:  24080502,
    4:  13551513,
    7:  61053501,
    11: 61053005,
    14: 24081002,
    20: 61950502,
    23: 23652007,
    30: 61055001,
    32: 61051501,
    33: 51350505,
    37: 42505002,
}

CUENTAS_REVERSION = {
    1:  {"gravado": 41750501, "no_gravado": 41750502},
    3:  24080597,
    4:  13551513,
    14: 24081075,
    20: 61970505,
    23: 23652091,
}

TC_ACTA       = {1, 2, 3, 4, 7, 11, 30, 32, 33, 37}
TC_COMISIONES = {20, 14, 23}
TC_EXCLUIR    = {0, 16, 40}

MESES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO",
    6: "JUNIO", 7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE",
    10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
}

NOMBRE_CUENTAS = {
    41550502: "Canon Gravado",       41550501: "Canon No Gravado",
    61053502: "Cuota Admon",         24080502: "IVA Comercial",
    13551513: "Retefte",             61053501: "EPM",
    61053005: "Seguro",              24081002: "IVA Comisión",
    61950502: "Comisión",            23652007: "Retefte Comisión",
    61055001: "Rep. Locativas",      61051501: "Prediales",
    51350505: "Aseo",                42505002: "Apropi. Depósito",
    41750501: "Rev. Canon Gravado",  41750502: "Rev. Canon Excluido",
    24080597: "Rev. IVA Comercial",  24081075: "Rev. IVA Comisión",
    61970505: "Rev. Comisión",       23652091: "Rev. Retefte Comisión",
    22050501: "Contrapartida Comisión",
    13050507: "Total Acta (contrapartida)",
}

NOMBRE_TC = {
    1: "Canon", 2: "Cuota Admon", 3: "IVA Comercial", 4: "Retefte",
    7: "EPM", 11: "Seguro", 14: "IVA Comisión", 20: "Comisión",
    23: "Retefte Comisión", 30: "Rep. Locativas", 32: "Prediales",
    33: "Aseo", 37: "Apropi. Depósito",
}


# ── Helpers ────────────────────────────────────────────────────────────────────
def parse_valor(val):
    if pd.isna(val):
        return 0.0
    clean = re.sub(r"[$ ,]", "", str(val)).strip()
    try:
        return float(clean)
    except ValueError:
        return 0.0


def es_reversion(detalle):
    return "revers" in str(detalle).lower() if pd.notna(detalle) else False


def format_fecha(fecha):
    if pd.isna(fecha):
        return ""
    if isinstance(fecha, str):
        parts = fecha.strip().split("/")
        if len(parts) == 3:
            return f"{parts[1]}/{parts[0]}/{parts[2]}"
        return fecha
    return fecha.strftime("%m/%d/%Y")


def get_mes(fecha):
    if pd.isna(fecha):
        return None
    if isinstance(fecha, str):
        parts = fecha.strip().split("/")
        if len(parts) == 3:
            try:
                return int(parts[1])
            except ValueError:
                return None
    try:
        return fecha.month
    except Exception:
        return None


def format_ad_ref(trans_int):
    """2026030005 → 'AD2630005'  (AD + año_corto + mes_sin_cero + secuencia)"""
    s = str(trans_int)
    if len(s) == 10:
        year_short = s[2:4]
        month = str(int(s[4:6]))
        seq = s[6:]
        return f"AD{year_short}{month}{seq}"
    return f"AD{trans_int}"


def get_cuenta(tc, reversion, tiene_iva):
    tabla = CUENTAS_REVERSION if reversion and tc in CUENTAS_REVERSION else CUENTAS
    entrada = tabla.get(tc)
    if entrada is None:
        return None
    if isinstance(entrada, dict):
        # Forward: CON IVA → No Gravado (41550501). SIN IVA → Gravado (41550502)
        # Reversal: CON IVA → Gravado reversal (41750501). SIN IVA → No Gravado reversal (41750502)
        if reversion:
            return entrada["gravado"] if tiene_iva else entrada["no_gravado"]
        else:
            return entrada["no_gravado"] if tiene_iva else entrada["gravado"]
    return entrada


def tipo_mov(debito, credito):
    """Crédito en fuente → 2 en planilla. Débito en fuente → 1."""
    return 2 if credito > 0 else 1


def make_row(cuenta, comp, fecha, doc, doc_ref, detalle, tipo, valor, base):
    return {
        "Cuenta":            cuenta,
        "Comprobante":       comp,
        "Fecha(mm/dd/yyyy)": fecha,
        "Documento":         doc,
        "Documento Ref.":    doc_ref,
        "Nit":               NIT_FONDO,
        "Detalle":           str(detalle) if pd.notna(detalle) else "",
        "Tipo":              tipo,
        "Valor":             valor,
        "Base":              base if pd.notna(base) else "",
        "Centro de Costo":   CENTRO_COSTO,
    }


# ── Procesador ACTA ────────────────────────────────────────────────────────────
def procesar_acta(df_src, documento, fecha_acta):
    """
    Un único Documento y una única fecha para todas las filas del ACTA.
    Siempre CP. Orden por No. Transaccion ASC, luego TipoCausa ASC.
    Gravado (41550502) = TC=1 SIN TC=3 paired.
    No Gravado (41550501) = TC=1 CON TC=3 paired.
    """
    trans_con_iva = set(
        df_src.loc[df_src["TipoCausa"] == 3, "No. Transaccion"]
        .dropna().astype(int).tolist()
    )

    filas = df_src[
        df_src["TipoCausa"].notna() &
        df_src["TipoCausa"].isin(TC_ACTA) &
        df_src["No. Transaccion"].notna()
    ].copy()
    filas["TipoCausa"] = filas["TipoCausa"].astype(int)
    filas["_trans"]    = filas["No. Transaccion"].astype(int)

    # Orden: primero TC=1/3/4 (Canon+IVA+Retefte) juntos por No.Trans ASC,
    # luego el resto de TCs por No.Trans ASC — dentro de cada grupo por TipoCausa ASC
    TC_GRUPO1 = {1, 3, 4}
    filas["_grupo"] = filas["TipoCausa"].apply(lambda t: 0 if t in TC_GRUPO1 else 1)
    filas = filas.sort_values(["_grupo", "_trans", "TipoCausa"], ascending=[True, True, True])

    resultado = []
    for _, row in filas.iterrows():
        tc       = int(row["TipoCausa"])
        detalle  = row["Detalle"]
        rev      = es_reversion(detalle)
        trans_id = int(row["_trans"])
        tiene_iva = trans_id in trans_con_iva

        cuenta = get_cuenta(tc, rev, tiene_iva)
        if cuenta is None:
            continue

        deb   = parse_valor(row["Debito"])
        cred  = parse_valor(row["Credito"])
        base  = parse_valor(row["Valor Base"])
        tipo  = tipo_mov(deb, cred)
        valor = cred if cred > 0 else deb

        resultado.append(make_row(
            cuenta, "CP", fecha_acta,      # siempre CP, siempre fecha_acta
            documento, documento,
            detalle, tipo, valor, base,
        ))

    # Fila final de contrapartida: "Total acta mes"
    # Tipo 1 = débito, Tipo 2 = crédito. Neto = total débitos − total créditos.
    # La contrapartida toma el tipo OPUESTO al mayor (para cuadrar la partida doble).
    if resultado:
        total_deb  = sum(r["Valor"] for r in resultado if r["Tipo"] == 1)
        total_cred = sum(r["Valor"] for r in resultado if r["Tipo"] == 2)
        neto = total_deb - total_cred
        # Si débitos > créditos → contrapartida es crédito (2); si créditos mayores → débito (1)
        tipo_total  = 2 if neto > 0 else 1
        valor_total = abs(neto)
        resultado.append(make_row(
            CUENTA_TOTAL_ACTA, "CP", fecha_acta,
            documento, documento,
            "Total acta mes", tipo_total, valor_total, "",
        ))

    return resultado


# ── Procesador COMISIONES ──────────────────────────────────────────────────────
def procesar_comisiones(df_src, cons_cp, cons_dv):
    """
    Agrupa por No. Transaccion. Cada grupo = una factura.
    Agrega fila de contrapartida 22050501 por grupo.
    CP (regulares) y DV (reversiones) tienen consecutivos separados.
    """
    filas = df_src[
        df_src["TipoCausa"].notna() &
        df_src["TipoCausa"].isin(TC_COMISIONES) &
        df_src["No. Transaccion"].notna()
    ].copy()
    filas["TipoCausa"] = filas["TipoCausa"].astype(int)
    filas["_rev"]      = filas["Detalle"].apply(es_reversion)

    resultado = []
    seq_cp = cons_cp
    seq_dv = cons_dv

    for trans_id, grupo in filas.groupby("No. Transaccion", sort=True):
        rev   = bool(grupo["_rev"].iloc[0])
        comp  = "DV" if rev else "CP"
        doc   = seq_dv if rev else seq_cp
        trans_int = int(trans_id)

        # Doc Ref: "FE" + No. Transaccion para CP; formato AD especial para DV
        doc_ref = format_ad_ref(trans_int) if rev else f"FE{trans_int}"

        fecha_str  = format_fecha(grupo.iloc[0]["Fecha"])
        mes        = get_mes(grupo.iloc[0]["Fecha"])
        mes_nombre = MESES.get(mes, "")

        sum_com = sum_iva = sum_ret = 0.0

        for _, row in grupo.iterrows():
            tc     = int(row["TipoCausa"])
            deb    = parse_valor(row["Debito"])
            cred   = parse_valor(row["Credito"])
            base   = parse_valor(row["Valor Base"]) if tc != 20 else ""  # TC=20 base vacío
            tipo   = tipo_mov(deb, cred)
            valor  = cred if cred > 0 else deb
            cuenta = get_cuenta(tc, rev, False)
            if cuenta is None:
                continue

            if tc == 20:
                sum_com += valor
            elif tc == 14:
                sum_iva += valor
            elif tc == 23:
                sum_ret += valor

            resultado.append(make_row(
                cuenta, comp, fecha_str, doc, doc_ref,
                row["Detalle"], tipo, valor, base,
            ))

        # Contrapartida 22050501
        valor_contra = sum_com + sum_iva - sum_ret
        # Tipo opuesto al de la comisión:
        # CP → TC=20 es DEBITO → tipo=1 → contra es 2
        # DV → TC=20 es CREDITO → tipo=2 → contra es 1
        tipo_contra = 1 if rev else 2
        resultado.append(make_row(
            CUENTA_CONTRA, comp, fecha_str, doc, doc_ref,
            f"COMISION MES {mes_nombre}", tipo_contra, valor_contra, "",
        ))

        if rev:
            seq_dv += 1
        else:
            seq_cp += 1

    return resultado


# ── Generar Excel ──────────────────────────────────────────────────────────────
COLS   = ["Cuenta", "Comprobante", "Fecha(mm/dd/yyyy)", "Documento",
          "Documento Ref.", "Nit", "Detalle", "Tipo", "Valor", "Base",
          "Centro de Costo"]
WIDTHS = [12, 13, 17, 13, 14, 13, 60, 6, 14, 10, 14]


def _hdr_acta(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font      = Font(bold=True, color=C_HEADER_FG, name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(bottom=Side(style="thin", color="000000"),
                                right=Side(style="thin", color="CCCCCC"))


def _fila(ws, row_num, ncols, alt):
    for c in range(1, ncols + 1):
        cell = ws.cell(row_num, c)
        cell.font      = Font(name="Arial", size=9)
        cell.alignment = Alignment(vertical="center")
        cell.border    = Border(bottom=Side(style="thin", color="E0E0E0"),
                                right=Side(style="thin", color="E0E0E0"))
        if alt:
            cell.fill = PatternFill("solid", fgColor=C_ALT)


def generar_excel(rows, sheet_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.freeze_panes = "A2"

    for c, name in enumerate(COLS, 1):
        ws.cell(1, c, name)
    _hdr_acta(ws, len(COLS))
    ws.row_dimensions[1].height = 30

    for i, row in enumerate(rows, 2):
        _fila(ws, i, len(COLS), i % 2 == 0)
        for c, key in enumerate(COLS, 1):
            v    = row.get(key, "")
            cell = ws.cell(i, c, v if v is not None else "")
            if c in (1, 4, 6):                                     # Cuenta, Documento, Nit
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c == 5:                                            # Documento Ref.
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c in (9, 10) and isinstance(v, (int, float)):     # Valor, Base
                cell.number_format = '#,##0;(#,##0);"-"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c in (2, 3, 8):                                    # Comprobante, Fecha, Tipo
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def resumen_cuentas(rows):
    df = pd.DataFrame(rows)
    res = (df.groupby("Cuenta")
             .agg(Filas=("Valor", "count"), Total=("Valor", "sum"))
             .reset_index())
    res["Nombre"] = res["Cuenta"].map(NOMBRE_CUENTAS).fillna("—")
    res["Total"]  = res["Total"].apply(lambda x: f"${x:,.0f}")
    return res[["Cuenta", "Nombre", "Filas", "Total"]]




import datetime

# ==============================================================================
# INTERFAZ — APP COMBINADA CONTABILIDAD RM2
# ==============================================================================

st.set_page_config(
    page_title="Contabilidad RM2",
    page_icon="🏢",
    layout="wide",
)

# ── CSS global ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

  /* Base */
  html, body, [data-testid="stAppViewContainer"] {
    font-family: 'Inter', sans-serif !important;
    background: #f4f6f9 !important;
  }
  [data-testid="stSidebar"] { display: none; }
  [data-testid="stMainBlockContainer"] { padding-top: 1.5rem !important; }

  /* Banner */
  .banner {
    background: linear-gradient(135deg, #0f2744 0%, #1a4a7a 60%, #2563ab 100%);
    border-radius: 16px;
    padding: 28px 40px;
    margin-bottom: 24px;
    display: flex;
    align-items: center;
    gap: 18px;
    box-shadow: 0 4px 20px rgba(15,39,68,.25);
  }
  .banner-logo {
    width: 52px; height: 52px; background: rgba(255,255,255,.15);
    border-radius: 12px; display: flex; align-items: center;
    justify-content: center; font-size: 28px; flex-shrink: 0;
  }
  .banner-title { font-size: 26px; font-weight: 800; color: #fff;
                  margin: 0; letter-spacing: -0.3px; }
  .banner-sub   { font-size: 13px; color: rgba(255,255,255,.6);
                  margin: 3px 0 0 0; font-weight: 400; }
  .banner-badge {
    margin-left: auto; background: rgba(255,255,255,.12);
    border: 1px solid rgba(255,255,255,.2); border-radius: 20px;
    padding: 4px 14px; font-size: 12px; color: rgba(255,255,255,.8);
    font-weight: 500; white-space: nowrap;
  }

  /* Navegación tipo segmented control */
  .nav-wrap {
    background: #fff; border-radius: 12px; padding: 6px;
    display: flex; gap: 4px; margin-bottom: 24px;
    box-shadow: 0 1px 4px rgba(0,0,0,.08);
  }
  .nav-btn {
    flex: 1; padding: 12px 20px; border-radius: 8px; border: none;
    cursor: pointer; font-family: 'Inter', sans-serif;
    font-size: 14px; font-weight: 600; transition: all .18s;
    display: flex; align-items: center; justify-content: center; gap: 8px;
  }
  .nav-btn.inactive { background: transparent; color: #6b7280; }
  .nav-btn.inactive:hover { background: #f3f4f6; color: #374151; }
  .nav-btn.active {
    background: linear-gradient(135deg, #0f2744, #1a4a7a);
    color: #fff; box-shadow: 0 2px 8px rgba(15,39,68,.3);
  }

  /* Tarjetas métricas */
  .kpi-row { display: flex; gap: 14px; margin-bottom: 22px; }
  .kpi {
    flex: 1; background: #fff; border-radius: 12px;
    padding: 18px 22px; box-shadow: 0 1px 4px rgba(0,0,0,.07);
    border-top: 3px solid #2563ab;
    transition: box-shadow .2s;
  }
  .kpi:hover { box-shadow: 0 4px 12px rgba(0,0,0,.1); }
  .kpi-label { font-size: 11px; font-weight: 600; color: #9ca3af;
               text-transform: uppercase; letter-spacing: .8px; }
  .kpi-value { font-size: 28px; font-weight: 800; color: #0f2744;
               margin: 6px 0 0 0; line-height: 1; }
  .kpi-sub   { font-size: 12px; color: #6b7280; margin: 4px 0 0 0; }

  /* Cards */
  .card {
    background: #fff; border-radius: 12px; padding: 24px;
    box-shadow: 0 1px 4px rgba(0,0,0,.07); margin-bottom: 18px;
  }
  .card-title {
    font-size: 11px; font-weight: 700; color: #6b7280;
    text-transform: uppercase; letter-spacing: .8px;
    margin-bottom: 16px; display: flex; align-items: center; gap: 6px;
  }

  /* Pills */
  .tag { display:inline-flex; align-items:center; gap:4px; padding:3px 10px;
         border-radius:6px; font-size:12px; font-weight:600; margin:2px; }
  .tag-blue  { background:#eff6ff; color:#1d4ed8; border:1px solid #bfdbfe; }
  .tag-green { background:#f0fdf4; color:#166534; border:1px solid #bbf7d0; }
  .tag-gray  { background:#f9fafb; color:#374151; border:1px solid #e5e7eb; }

  /* Tabs */
  [data-testid="stTabs"] > div > div { gap: 4px; border-bottom: 2px solid #e5e7eb !important; }
  [data-testid="stTab"] {
    border-radius: 8px 8px 0 0 !important;
    font-weight: 600 !important; font-size: 13px !important;
    color: #6b7280 !important; padding: 10px 18px !important;
    border-bottom: 2px solid transparent !important;
  }
  [data-testid="stTab"][aria-selected="true"] {
    color: #0f2744 !important;
    border-bottom: 2px solid #0f2744 !important;
    background: transparent !important;
  }
  [data-testid="stTab"] p, [data-testid="stTab"] span { color: inherit !important; }

  /* Botones primarios Streamlit */
  [data-testid="stButton"] > button[kind="primary"] {
    background: linear-gradient(135deg, #0f2744, #1a4a7a) !important;
    color: #fff !important; border: none !important; border-radius: 8px !important;
    font-weight: 600 !important; font-size: 14px !important;
    letter-spacing: .2px !important; box-shadow: 0 2px 8px rgba(15,39,68,.25) !important;
    transition: all .2s !important;
  }
  [data-testid="stButton"] > button[kind="primary"]:hover {
    background: linear-gradient(135deg, #0a1e35, #0f2744) !important;
    box-shadow: 0 4px 12px rgba(15,39,68,.35) !important;
    transform: translateY(-1px) !important;
  }
  [data-testid="stButton"] > button[kind="secondary"] {
    background: #fff !important; color: #374151 !important;
    border: 1.5px solid #d1d5db !important; border-radius: 8px !important;
    font-weight: 600 !important; font-size: 14px !important;
  }
  [data-testid="stButton"] > button[kind="secondary"]:hover {
    border-color: #0f2744 !important; color: #0f2744 !important;
    background: #f8faff !important;
  }

  /* Botón descarga */
  [data-testid="stDownloadButton"] > button {
    background: linear-gradient(135deg, #0f2744, #1a4a7a) !important;
    color: #fff !important; border: none !important; border-radius: 8px !important;
    font-weight: 600 !important; font-size: 14px !important;
    box-shadow: 0 2px 8px rgba(15,39,68,.25) !important; padding: 12px !important;
  }
  [data-testid="stDownloadButton"] > button:hover {
    background: linear-gradient(135deg, #0a1e35, #0f2744) !important;
    box-shadow: 0 4px 12px rgba(15,39,68,.35) !important;
    transform: translateY(-1px) !important;
  }

  /* Inputs */
  [data-testid="stNumberInput"] label,
  [data-testid="stDateInput"]   label,
  [data-testid="stFileUploader"] label {
    color: #374151 !important; font-weight: 500 !important; font-size: 13px !important;
  }
  [data-testid="stFileUploader"] {
    border: 2px dashed #d1d5db !important; border-radius: 10px !important;
    background: #fafafa !important;
  }

  /* Divider */
  hr { border-color: #e5e7eb !important; margin: 20px 0 !important; }

  /* Headings */
  h3 { color: #0f2744 !important; font-weight: 700 !important; font-size: 18px !important; }
</style>
""", unsafe_allow_html=True)

# ── Banner ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="banner">
  <div class="banner-logo">🏢</div>
  <div>
    <p class="banner-title">Contabilidad RM2</p>
    <p class="banner-sub">Herramientas de conciliación y generación de plantillas contables</p>
  </div>
  <div class="banner-badge">RM2 · Fondo Raiz</div>
</div>
""", unsafe_allow_html=True)

# ── Navegación ─────────────────────────────────────────────────────────────────
if "modulo" not in st.session_state:
    st.session_state["modulo"] = None

col_n1, col_n2 = st.columns(2)
with col_n1:
    active = st.session_state["modulo"] == "conciliacion"
    if st.button("🔍  Conciliación", use_container_width=True,
                 type="primary" if active else "secondary", key="nav_concil"):
        st.session_state["modulo"] = "conciliacion"
        st.rerun()
with col_n2:
    active = st.session_state["modulo"] == "contai"
    if st.button("📋  Actas Fondo — Contai", use_container_width=True,
                 type="primary" if active else "secondary", key="nav_contai"):
        st.session_state["modulo"] = "contai"
        st.rerun()

modulo = st.session_state["modulo"]

# ── Landing ────────────────────────────────────────────────────────────────────
if modulo is None:
    st.markdown("""
    <div style="padding:48px 0 32px; text-align:center;">
      <p style="font-size:40px;margin:0;">👋</p>
      <p style="font-size:20px;font-weight:700;color:#0f2744;margin:14px 0 6px;">
        Bienvenido a Contabilidad RM2
      </p>
      <p style="font-size:14px;color:#6b7280;margin:0;">
        Selecciona un módulo arriba para comenzar
      </p>
    </div>
    <div style="display:flex;gap:20px;max-width:680px;margin:0 auto 48px;">
      <div style="flex:1;background:#fff;border-radius:14px;padding:28px;
                  box-shadow:0 2px 12px rgba(0,0,0,.08);border-top:4px solid #0f2744;">
        <div style="font-size:28px;margin-bottom:12px;">🔍</div>
        <p style="font-size:15px;font-weight:700;color:#0f2744;margin:0 0 8px;">Conciliación</p>
        <p style="font-size:13px;color:#6b7280;margin:0;line-height:1.6;">
          Verifica propietarios, tipo causa, IVA y retenciones cruzando el archivo
          de propietarios con contabilidad.
        </p>
      </div>
      <div style="flex:1;background:#fff;border-radius:14px;padding:28px;
                  box-shadow:0 2px 12px rgba(0,0,0,.08);border-top:4px solid #2563ab;">
        <div style="font-size:28px;margin-bottom:12px;">📋</div>
        <p style="font-size:15px;font-weight:700;color:#0f2744;margin:0 0 8px;">Actas Fondo — Contai</p>
        <p style="font-size:13px;color:#6b7280;margin:0;line-height:1.6;">
          Genera las planillas contables de ACTA y COMISIONES desde el
          Movimiento Cta 28150501.
        </p>
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO: CONCILIACIÓN
# ══════════════════════════════════════════════════════════════════════════════
if modulo == "conciliacion":
    st.markdown("### 🔍 Conciliación RM2")
    st.markdown("Sube los archivos y haz clic en **Conciliar** para generar el reporte.")

    with st.expander("📥 ¿Cómo descargar los archivos del sistema?"):
        st.markdown(
            """
### 1. Auxiliar 28150501

**Ruta:** `Módulo Contabilidad` → `Análisis e Informes` → `Auxiliares`

1. Activa el rango de fechas: **Fecha inicial** y **Fecha final** del periodo a conciliar.
2. En **Cuenta**, ingresa: **`28150501`**
3. Activa **todos los filtros (columnas)** disponibles para que el reporte
   incluya toda la información necesaria (NIT, nombre tercero, detalle,
   débito, crédito, número de inmueble, tipo de causa, etc.).
4. Genera y exporta el reporte a Excel (`.xls` o `.xlsx`).

> ℹ️ Asegúrate de exportar con **todas las columnas activas**; si falta alguna,
> la conciliación puede arrojar resultados incompletos.

### 2. Informe actualizado de Propietarios

Genera el informe **actualizado** de propietarios.

**Ruta:** `Módulo de Arriendos` → `Elaboración de Contratos`

1. **Activa todos** los filtros / columnas disponibles.
2. **Envía a Excel** para exportar el reporte (`.xls` o `.xlsx`).
            """
        )

    col1, col2 = st.columns(2)
    with col1:
        f_prop = st.file_uploader("📋 Propietarios RM2", type=["xls","xlsx"],
                                  help="Informe actualizado de Propietarios (todas las columnas)")
    with col2:
        f_cont = st.file_uploader("📒 Auxiliar 28150501", type=["xls","xlsx"],
                                  help="Auxiliar original de la cuenta 28150501 (todas las columnas)")

    # ── Tipo Causa: catálogo establecido (opcional actualizar) ──────────────────
    f_tc = None
    actualizar_tc = st.checkbox(
        "🔄 Actualizar catálogo de Tipo Causa",
        help="Por defecto se usa el catálogo establecido. Actívalo solo si los Tipo Causa cambiaron."
    )
    if actualizar_tc:
        f_tc = st.file_uploader("📄 Tipo Causa (nuevo catálogo)", type=["xls","xlsx"],
                                help="Sube el archivo TIPO CAUSA actualizado")
        if f_tc is None:
            st.warning("Sube el archivo de Tipo Causa o desactiva la casilla para usar el catálogo establecido.")
    else:
        st.caption(f"✅ Usando el catálogo de Tipo Causa establecido ({len(DEFAULT_TIPO_CAUSA)} conceptos).")
        with st.expander("Ver catálogo establecido"):
            st.dataframe(
                pd.DataFrame(
                    [{"Tipo Causa": k, "Concepto Contabilidad": v} for k, v in DEFAULT_TIPO_CAUSA.items()]
                ),
                hide_index=True, use_container_width=True
            )

    st.divider()

    if st.button("▶ Conciliar", type="primary", disabled=not (f_prop and f_cont)):
        with st.spinner("Procesando..."):
            df_v1, df_v2, df_v3, crosscheck, analisis, prop_inm, cont_inm = run_conciliacion(f_prop, f_cont, f_tc)

        total_inm = df_v1["Inmueble"].nunique()
        correctos = len(df_v1[df_v1["Estado"]=="CORRECTO"])
        alertas   = len(df_v1[df_v1["Estado"].str.startswith("ALERTA")])
        solo_p    = len(df_v1[df_v1["Estado"]=="SOLO EN PROPIETARIOS"])
        solo_c    = len(df_v1[df_v1["Estado"]=="SOLO EN CONTABILIDAD"])
        total_v3  = len(df_v3)
        iva_v3    = len(df_v3[df_v3["tipo_causa"]=="3"])
        rete_v3   = len(df_v3[df_v3["tipo_causa"]=="4"])
        cc        = crosscheck

        st.subheader("Resumen")
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("Inmuebles verificados", total_inm)
        m2.metric("✅ Correctos", correctos)
        m3.metric("🚨 Alertas", alertas, delta=alertas if alertas else None, delta_color="inverse")
        m4.metric("Solo Propietarios", solo_p)
        m5.metric("Solo Contabilidad", solo_c)

        c1, c2 = st.columns(2)
        with c1:
            if cc["ok_iva"]:
                st.success(f"✅ IVA cuadra — TC20 × 19% = TC14 ({cc['iva_calc']:,})")
            else:
                st.error(f"❌ IVA NO cuadra — Calculado: {cc['iva_calc']:,} · Real (TC14): {cc['neto_14']:,} · **Diferencia: {cc['dif_iva']:,}**")
        with c2:
            if cc["ok_rete"]:
                st.success(f"✅ Retefuente cuadra — TC20 × 11% = TC23 ({cc['rete_calc']:,})")
            else:
                st.error(f"❌ Retefuente NO cuadra — Calculado: {cc['rete_calc']:,} · Real (TC23): {cc['neto_23']:,} · **Diferencia: {cc['dif_rete']:,}**")

        st.divider()

        tab1, tab2, tab3 = st.tabs([
            "📋 Verificación 1 — Propietarios",
            "📄 Verificación 2 — Tipo Causa",
            "💰 Verificación 3 — IVA y Retenciones",
        ])

        with tab1:
            st.markdown("**Resultado por inmueble** (solo propietarios activos `Ex Prop = False`)")
            def color_estado_v1(val):
                if val == "CORRECTO":               return "background-color:#C6EFCE;color:#276221"
                if "ALERTA" in str(val):            return "background-color:#FFCCCC;color:#9C0006"
                if val == "SOLO EN PROPIETARIOS":   return "background-color:#DAE3F3;color:#1F4E79"
                if val == "SOLO EN CONTABILIDAD":   return "background-color:#FCE4D6;color:#843C0C"
                if val == "PROPIETARIO ADICIONAL":  return "background-color:#FFEB9C;color:#9C6500"
                return ""
            display_v1 = df_v1.rename(columns={"Inmueble":"No. Inm","Propietario_Prop":"Propietario (Prop.)","NIT_Prop":"NIT (Prop.)","Propietario_Cont":"Propietario (Cont.)","NIT_Cont":"NIT (Cont.)"})
            filtro = st.selectbox("Filtrar por estado", ["Todos","CORRECTO","ALERTA — NOMBRE DIFIERE","ALERTA — NIT NO COINCIDE","SOLO EN PROPIETARIOS","SOLO EN CONTABILIDAD","PROPIETARIO ADICIONAL"])
            if filtro != "Todos":
                display_v1 = display_v1[display_v1["Estado"] == filtro]
            st.dataframe(display_v1.style.map(color_estado_v1, subset=["Estado"]), use_container_width=True, hide_index=True, height=500)

        with tab2:
            st.markdown("**Resultado por fila** de Contabilidad (ordenado por inmueble)")
            def color_estado_v2(val):
                if val == "CORRECTO":                    return "background-color:#C6EFCE;color:#276221"
                if val == "REVISAR":                     return "background-color:#FFEB9C;color:#9C6500"
                if val in ("CÓDIGO NO EXISTE","SIN KEYWORDS"): return "background-color:#FFCCCC;color:#9C0006"
                return ""
            display_v2 = df_v2.rename(columns={"Inmueble":"No. Inm","Concepto":"Concepto (Catálogo)","Detalle":"Detalle Contabilidad","Observacion":"Observación"})
            filtro2 = st.selectbox("Filtrar por estado ", ["Todos","CORRECTO","REVISAR","CÓDIGO NO EXISTE"])
            if filtro2 != "Todos":
                display_v2 = display_v2[display_v2["Estado"] == filtro2]
            st.dataframe(display_v2.style.map(color_estado_v2, subset=["Estado"]), use_container_width=True, hide_index=True, height=500)

        with tab3:
            st.markdown("#### Canon de arrendamiento (TC 1) — neto (débito − crédito) por tipo de inmueble")
            cols_tc1 = st.columns(len(crosscheck["tc1_by_type"]) + 1)
            for col, d in zip(cols_tc1, crosscheck["tc1_by_type"]):
                col.metric(f"{d['tipo']} ({d['cantidad']})", f"{d['neto']:,}")
            cols_tc1[-1].metric(f"TOTAL TC1 ({crosscheck['tc1_cant_total']})",
                                f"{crosscheck['tc1_neto_total']:,}")
            df_tc1 = pd.DataFrame(crosscheck["tc1_by_type"]).rename(columns={
                "etiqueta":"Tipo de inmueble", "cantidad":"Cantidad", "neto":"Neto (déb − créd)"
            })[["Tipo de inmueble","Cantidad","Neto (déb − créd)"]]
            st.dataframe(df_tc1, use_container_width=True, hide_index=True)

            df_sc = crosscheck["tc1_sin_clasif"]
            if len(df_sc):
                st.markdown(f"##### ⚠ Canon (TC 1) sin clasificación — {len(df_sc)} registro(s)")
                st.caption("No se reconoció el código de inmueble (AP / LC / PQV / PQM / OFC / C.U) en el detalle.")
                disp_sc = df_sc.rename(columns={
                    "inm":"No. Inm", "nit":"NIT", "nombre_cont":"Nombre Tercero",
                    "detalle":"Detalle", "neto":"Neto (déb − créd)",
                })[["No. Inm","NIT","Nombre Tercero","Detalle","Neto (déb − créd)"]]
                st.dataframe(disp_sc, use_container_width=True, hide_index=True)
            else:
                st.success("✓ Todos los canon (TC 1) quedaron clasificados.")

            st.divider()
            st.markdown("#### IVA y Retenciones — neto (débito − crédito)")
            st.markdown(f"**IVA Arrendamiento (TC 3): {iva_v3} registros · Retención Fuente Canon (TC 4): {rete_v3} registros**")
            fa, fb = st.columns(2)
            fa.metric("Neto TC 3 — IVA (déb − créd)", f"{crosscheck['tc3_neto']:,}")
            fb.metric("Neto TC 4 — Retención (déb − créd)", f"{crosscheck['tc4_neto']:,}")
            filtro3 = st.selectbox("Filtrar por tipo causa", ["Todos","3 — IVA Arrendamiento","4 — Retención en la Fuente"])
            display_v3 = df_v3.rename(columns={"inm_num":"No. Inm","nombre_cont":"Nombre Tercero","tipo_causa":"TC","concepto":"Concepto","detalle":"Detalle","debito":"Débito","credito":"Crédito","neto":"Neto"})[["No. Inm","Nombre Tercero","TC","Concepto","Detalle","Débito","Crédito","Neto"]]
            if filtro3 == "3 — IVA Arrendamiento":      display_v3 = display_v3[display_v3["TC"]=="3"]
            elif filtro3 == "4 — Retención en la Fuente": display_v3 = display_v3[display_v3["TC"]=="4"]
            def color_tc(val):
                if val == "3": return "background-color:#DEEAF1;color:#1F4E79"
                if val == "4": return "background-color:#EBF0DE;color:#375623"
                return ""
            def color_neto(val):
                if isinstance(val,(int,float)) and val < 0: return "color:#9C0006;font-weight:bold"
                return ""
            st.dataframe(display_v3.style.map(color_tc,subset=["TC"]).map(color_neto,subset=["Neto"]), use_container_width=True, hide_index=True, height=500)

        st.divider()
        excel_buf = build_excel(df_v1, df_v2, df_v3, crosscheck, analisis)
        st.download_button("⬇️ Descargar REPORTE_CONCILIACION.xlsx", data=excel_buf,
                           file_name="REPORTE_CONCILIACION.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           type="primary")
    else:
        if not (f_prop and f_cont):
            st.info("👆 Sube **Propietarios** y **Contabilidad** para habilitar el botón **Conciliar**. *Tipo Causa* es opcional.")

# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO: ACTAS FONDO — CONTAI
# ══════════════════════════════════════════════════════════════════════════════
elif modulo == "contai":
    st.markdown("### 📋 Plantilla Contai")

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p class="card-title">📂 Archivo de movimientos</p>', unsafe_allow_html=True)
    archivo = st.file_uploader(
        "Arrastra o selecciona · Movimiento Cta 28150501 RM2 -Fondo.xls",
        type=["xls","xlsx"],
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if not archivo:
        st.markdown("""
        <div style="text-align:center;padding:40px 0;color:#9ca3af;">
          <div style="font-size:40px">📄</div>
          <p style="font-size:15px;margin-top:10px;">Carga el archivo de movimientos para comenzar</p>
        </div>""", unsafe_allow_html=True)
        st.stop()

    try:
        df_src = pd.read_excel(archivo, sheet_name=0, header=0)
    except Exception as e:
        st.error(f"Error al leer el archivo: {e}")
        st.stop()

    df_src["TipoCausa"] = pd.to_numeric(df_src["TipoCausa"], errors="coerce")

    # ── Selección de propietario (NIT / Nombre Tercero) ──────────────────────────
    df_owners = df_src[df_src["Nit"].notna() & df_src["No. Transaccion"].notna()].copy()
    df_owners["Nit"] = pd.to_numeric(df_owners["Nit"], errors="coerce")
    owners = (df_owners.dropna(subset=["Nit"])
              .groupby("Nit")["Nombre Tercero"].first().reset_index())
    opciones = ["Todos los propietarios"] + [
        f"{r['Nombre Tercero']} ({int(r['Nit'])})" for _, r in owners.iterrows()
    ]
    nit_por_opcion = {
        f"{r['Nombre Tercero']} ({int(r['Nit'])})": int(r["Nit"])
        for _, r in owners.iterrows()
    }

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p class="card-title">👤 Propietario</p>', unsafe_allow_html=True)
    sel_prop = st.selectbox(
        "Propietario al que se le genera la plantilla",
        opciones,
        help="Filtra los movimientos por NIT / Nombre Tercero antes de generar las planillas",
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if sel_prop != "Todos los propietarios":
        nit_sel = nit_por_opcion[sel_prop]
        df_src = df_src[pd.to_numeric(df_src["Nit"], errors="coerce") == nit_sel].copy()

    n_acta = len(df_src[df_src["TipoCausa"].isin(TC_ACTA) & df_src["No. Transaccion"].notna()])
    n_com  = len(df_src[df_src["TipoCausa"].isin(TC_COMISIONES) & df_src["No. Transaccion"].notna()])
    n_fact = df_src[df_src["TipoCausa"].isin(TC_COMISIONES) & df_src["No. Transaccion"].notna()]["No. Transaccion"].nunique()

    st.markdown(f"""
    <div class="kpi-row">
      <div class="kpi" style="border-top-color:#0f2744;">
        <div class="kpi-label">Archivo</div>
        <div class="kpi-value" style="font-size:14px;margin-top:8px;font-weight:600;">✅ {archivo.name}</div>
      </div>
      <div class="kpi" style="border-top-color:#2563ab;">
        <div class="kpi-label">Filas ACTA</div>
        <div class="kpi-value">{n_acta:,}</div>
      </div>
      <div class="kpi" style="border-top-color:#16a34a;">
        <div class="kpi-label">Filas Comisiones</div>
        <div class="kpi-value">{n_com:,}</div>
      </div>
      <div class="kpi" style="border-top-color:#d97706;">
        <div class="kpi-label">Facturas únicas</div>
        <div class="kpi-value">{n_fact:,}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    tab_acta, tab_com = st.tabs(["📋  Planilla ACTA", "💼  Planilla COMISIONES"])

    with tab_acta:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<p class="card-title">⚙️ Parámetros del ACTA</p>', unsafe_allow_html=True)
        st.markdown('<p style="font-size:13px;color:#6b7280;margin-bottom:16px;">El mismo Documento y Fecha se aplican a <strong>todas</strong> las filas. Canon va primero, Prediales/Admon al final.</p>', unsafe_allow_html=True)
        col_a1, col_a2 = st.columns(2)
        with col_a1:
            doc_acta = st.number_input("Consecutivo Documento", min_value=1, value=26040185, step=1, key="doc_acta")
        with col_a2:
            fecha_acta_dt = st.date_input("Fecha del ACTA", value=datetime.date(2026, 4, 30), key="fecha_acta")
        fecha_acta_str = fecha_acta_dt.strftime("%m/%d/%Y")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div style="display:flex;flex-wrap:wrap;gap:6px;margin-bottom:14px;"><span class="tag tag-blue">Canon · IVA · Retefte</span><span class="tag tag-blue">Admon · EPM · Seguro</span><span class="tag tag-blue">Prediales · Aseo · Rep. Locativas</span><span class="tag tag-green">Siempre CP</span></div>', unsafe_allow_html=True)

        if st.button("▶  Generar Planilla ACTA", type="primary", key="btn_acta", use_container_width=True):
            with st.spinner("Procesando ACTA…"):
                rows_acta = procesar_acta(df_src, int(doc_acta), fecha_acta_str)
            if not rows_acta:
                st.warning("No se encontraron movimientos válidos para el ACTA.")
            else:
                st.success(f"✅ **{len(rows_acta):,} filas** generadas — Documento: **{doc_acta}** · Fecha: **{fecha_acta_str}**")
                col_r1, col_r2 = st.columns([1, 2])
                with col_r1:
                    st.markdown("**Resumen por cuenta**")
                    st.dataframe(resumen_cuentas(rows_acta), use_container_width=True, hide_index=True, height=320)
                with col_r2:
                    with st.expander("Vista previa — primeras 50 filas", expanded=True):
                        cols_show = [c for c in pd.DataFrame(rows_acta).columns if not c.startswith("_")]
                        st.dataframe(pd.DataFrame(rows_acta)[cols_show].head(50), use_container_width=True)
                st.markdown("---")
                buf = generar_excel(rows_acta, "ACTA_RM2")
                st.download_button("⬇️  Descargar Plantilla Acta RM2.xlsx", data=buf,
                                   file_name="Plantilla_Acta_RM2.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True, type="primary", key="dl_acta")

    with tab_com:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<p class="card-title">⚙️ Parámetros de COMISIONES</p>', unsafe_allow_html=True)
        st.markdown('<p style="font-size:13px;color:#6b7280;margin-bottom:16px;">Cada factura recibe su propio consecutivo. Contrapartida <strong>22050501</strong> automática. Reversiones usan comprobante <strong>DV</strong>.</p>', unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        with col1:
            cons_cp = st.number_input("Consecutivo inicial CP (regulares)", min_value=1, value=26040186, step=1, key="cons_cp")
        with col2:
            cons_dv = st.number_input("Consecutivo inicial DV (reversiones)", min_value=1, value=26040001, step=1, key="cons_dv")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div style="display:flex;flex-wrap:wrap;gap:6px;margin-bottom:14px;"><span class="tag tag-blue">Comisión · IVA Comisión · Retefte Comisión</span><span class="tag tag-gray">Contrapartida 22050501 automática</span><span class="tag tag-green">CP regulares · DV reversiones</span></div>', unsafe_allow_html=True)

        if st.button("▶  Generar Planilla COMISIONES", type="primary", key="btn_com", use_container_width=True):
            with st.spinner("Procesando Comisiones…"):
                rows_com = procesar_comisiones(df_src, int(cons_cp), int(cons_dv))
            if not rows_com:
                st.warning("No se encontraron movimientos de comisiones.")
            else:
                n_cp = sum(1 for r in rows_com if r["Comprobante"] == "CP")
                n_dv = sum(1 for r in rows_com if r["Comprobante"] == "DV")
                st.success(f"✅ **{len(rows_com):,} filas** generadas — CP: **{n_cp}** · DV: **{n_dv}** · Contrapartidas incluidas")
                col_r1, col_r2 = st.columns([1, 2])
                with col_r1:
                    st.markdown("**Resumen por cuenta**")
                    st.dataframe(resumen_cuentas(rows_com), use_container_width=True, hide_index=True, height=320)
                with col_r2:
                    with st.expander("Vista previa — primeras 60 filas", expanded=True):
                        cols_show = [c for c in pd.DataFrame(rows_com).columns if not c.startswith("_")]
                        st.dataframe(pd.DataFrame(rows_com)[cols_show].head(60), use_container_width=True)
                st.markdown("---")
                buf = generar_excel(rows_com, "COMISIONES_RM2")
                st.download_button("⬇️  Descargar Plantilla Comisiones RM2.xlsx", data=buf,
                                   file_name="Plantilla_Comisiones_RM2.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True, type="primary", key="dl_com")
